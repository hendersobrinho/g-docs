from __future__ import annotations

from datetime import date
import re
import unicodedata

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font
except ImportError:  # pragma: no cover - depende do ambiente local
    Workbook = None
    Font = None
    load_workbook = None

from documentos_empresa_app.services.documento_service import DocumentoService
from documentos_empresa_app.services.empresa_service import EmpresaService
from documentos_empresa_app.services.periodo_service import PeriodoService
from documentos_empresa_app.services.status_service import StatusService
from documentos_empresa_app.services.tipo_service import TipoService
from documentos_empresa_app.utils.common import (
    AUTO_STATUS_NAO_COBRAR,
    MONTH_NAMES,
    ValidationError,
)


EMPRESA_IMPORT_LAYOUT = (
    {
        "index": 1,
        "field": "codigo_empresa",
        "label": "Codigo da empresa",
        "required": True,
        "example": "101",
    },
    {
        "index": 2,
        "field": "nome_empresa",
        "label": "Nome da empresa",
        "required": True,
        "example": "Empresa Exemplo Ltda",
    },
    {
        "index": 3,
        "field": "email_contato",
        "label": "Email de contato",
        "required": False,
        "example": "contato@empresa.com",
    },
    {
        "index": 4,
        "field": "nome_contato",
        "label": "Nome do contato",
        "required": False,
        "example": "Maria Silva",
    },
    {
        "index": 5,
        "field": "observacao",
        "label": "Observacao",
        "required": False,
        "example": "Prefere retorno ate o dia 10",
    },
)

EMPRESA_IMPORT_HEADER_ALIASES = {
    0: {"codigo", "codigo_empresa", "codigo_da_empresa"},
    1: {"nome", "nome_empresa", "nome_da_empresa"},
    2: {"email", "email_contato", "email_para_contato"},
    3: {"contato", "nome_contato", "nome_do_contato"},
    4: {"observacao", "obs", "anotacao", "anotacoes", "observacoes"},
}

DOCUMENTO_IMPORT_LAYOUT = (
    {
        "index": 1,
        "field": "meios_recebimento",
        "label": "Meios de recebimento",
        "required": False,
        "example": "Email, Onvio",
    },
    {
        "index": 2,
        "field": "nome_documento",
        "label": "Nome do documento",
        "required": True,
        "example": "Banco do Brasil",
    },
    {
        "index": 3,
        "field": "nome_tipo",
        "label": "Tipo do documento",
        "required": True,
        "example": "Extratos CC",
    },
)

DOCUMENTO_IMPORT_HEADER_ALIASES = {
    0: {"meios", "meio", "meios_recebimento", "meios_de_recebimento", "meio_de_recebimento"},
    1: {"nome_documento", "documento", "nome", "nome_do_documento"},
    2: {"tipo", "tipos", "nome_tipo", "tipo_documento", "tipo_do_documento"},
}

LEGACY_DOCUMENTO_IMPORT_HEADER_ALIASES = {
    0: {"nome_documento", "documento", "nome", "nome_do_documento"},
    1: {"tipo", "tipos", "nome_tipo", "tipo_documento", "tipo_do_documento"},
}

CADASTRO_COMPLETO_FIXED_LAYOUT = (
    {
        "index": 1,
        "field": "codigo_empresa",
        "label": "Codigo da empresa",
        "required": True,
        "example": "101",
    },
    {
        "index": 2,
        "field": "nome_empresa",
        "label": "Nome da empresa",
        "required": True,
        "example": "Empresa Exemplo Ltda",
    },
    {
        "index": 3,
        "field": "email_contato",
        "label": "Email de contato",
        "required": False,
        "example": "contato@empresa.com",
    },
    {
        "index": 4,
        "field": "nome_contato",
        "label": "Nome do contato",
        "required": False,
        "example": "Maria Silva",
    },
    {
        "index": 5,
        "field": "nome_documento",
        "label": "Nome do documento",
        "required": False,
        "example": "Banco do Brasil",
    },
    {
        "index": 6,
        "field": "meios_recebimento",
        "label": "Meio de recebimento",
        "required": False,
        "example": "Email, Onvio",
    },
    {
        "index": 7,
        "field": "nome_tipo",
        "label": "Tipo do documento",
        "required": False,
        "example": "Extratos CC",
    },
    {
        "index": 8,
        "field": "observacao",
        "label": "Observacao",
        "required": False,
        "example": "Prefere retorno ate o dia 10",
    },
)

CADASTRO_COMPLETO_MONTH_LAYOUT = tuple(
    {
        "index": month_number + 8,
        "field": f"status_{month_number:02d}",
        "label": month_name,
        "required": False,
        "example": {1: "OK", 2: "P", 3: "X"}.get(month_number, ""),
    }
    for month_number, month_name in MONTH_NAMES.items()
)

CADASTRO_COMPLETO_IMPORT_LAYOUT = CADASTRO_COMPLETO_FIXED_LAYOUT + CADASTRO_COMPLETO_MONTH_LAYOUT

CADASTRO_COMPLETO_FIELD_ALIASES = {
    "codigo_empresa": set(EMPRESA_IMPORT_HEADER_ALIASES[0]),
    "nome_empresa": set(EMPRESA_IMPORT_HEADER_ALIASES[1]),
    "email_contato": set(EMPRESA_IMPORT_HEADER_ALIASES[2]),
    "nome_contato": set(EMPRESA_IMPORT_HEADER_ALIASES[3]),
    "nome_documento": set(DOCUMENTO_IMPORT_HEADER_ALIASES[1]),
    "meios_recebimento": set(DOCUMENTO_IMPORT_HEADER_ALIASES[0]) | {"forma_recebimento", "forma_de_recebimento"},
    "nome_tipo": set(DOCUMENTO_IMPORT_HEADER_ALIASES[2]),
    "observacao": set(EMPRESA_IMPORT_HEADER_ALIASES[4]),
}

CADASTRO_COMPLETO_LEGACY_FIELD_INDEXES = {
    "codigo_empresa": 0,
    "nome_empresa": 1,
    "email_contato": 2,
    "nome_contato": 3,
    "meios_recebimento": 4,
    "nome_documento": 5,
    "nome_tipo": 6,
    "observacao": 7,
}

MONTH_HEADER_TO_NUMBER = {
    "janeiro": 1,
    "fevereiro": 2,
    "marco": 3,
    "abril": 4,
    "maio": 5,
    "junho": 6,
    "julho": 7,
    "agosto": 8,
    "setembro": 9,
    "outubro": 10,
    "novembro": 11,
    "dezembro": 12,
}

IMPORT_STATUS_ALIASES = {
    "ok": "Recebido",
    "recebido": "Recebido",
    "p": "Pendente",
    "pendente": "Pendente",
    "x": AUTO_STATUS_NAO_COBRAR,
    "nao_possui": AUTO_STATUS_NAO_COBRAR,
    "nao_tem": AUTO_STATUS_NAO_COBRAR,
    "nao_cobrar": AUTO_STATUS_NAO_COBRAR,
    "sem_cobranca": AUTO_STATUS_NAO_COBRAR,
    "encerrado": "Encerrado",
}

IMPORT_COMPANY_CREATED = "created"
IMPORT_COMPANY_UPDATED = "updated"
IMPORT_COMPANY_REUSED = "reused"
IMPORT_COMPANY_OUTCOME_PRIORITY = {
    IMPORT_COMPANY_REUSED: 1,
    IMPORT_COMPANY_UPDATED: 2,
    IMPORT_COMPANY_CREATED: 3,
}


class ImportService:
    def __init__(
        self,
        empresa_service: EmpresaService,
        tipo_service: TipoService,
        documento_service: DocumentoService,
        periodo_service: PeriodoService,
        status_service: StatusService,
    ) -> None:
        self.empresa_service = empresa_service
        self.tipo_service = tipo_service
        self.documento_service = documento_service
        self.periodo_service = periodo_service
        self.status_service = status_service

    def get_empresa_import_layout(self) -> tuple[dict, ...]:
        return EMPRESA_IMPORT_LAYOUT

    def get_documento_import_layout(self) -> tuple[dict, ...]:
        return DOCUMENTO_IMPORT_LAYOUT

    def get_cadastro_completo_import_layout(self) -> tuple[dict, ...]:
        return CADASTRO_COMPLETO_IMPORT_LAYOUT

    def export_empresa_template(self, file_path: str) -> dict:
        return self._export_template(
            file_path,
            worksheet_title="Empresas",
            layout=EMPRESA_IMPORT_LAYOUT,
        )

    def export_documento_template(self, file_path: str) -> dict:
        return self._export_template(
            file_path,
            worksheet_title="Documentos",
            layout=DOCUMENTO_IMPORT_LAYOUT,
        )

    def export_cadastro_completo_template(self, file_path: str) -> dict:
        return self._export_template(
            file_path,
            worksheet_title="Cadastros completos",
            layout=CADASTRO_COMPLETO_IMPORT_LAYOUT,
        )

    def import_empresas(self, file_path: str) -> dict:
        workbook = self._load_workbook(file_path)
        worksheet = workbook.active

        imported = 0
        failed = 0
        errors: list[str] = []

        for row_number, row in self._iter_data_rows(worksheet, self._looks_like_empresa_header):
            try:
                self._import_empresa_row(row)
                imported += 1
            except ValidationError as exc:
                failed += 1
                errors.append(f"Linha {row_number}: {exc}")

        return {"imported": imported, "failed": failed, "errors": errors}

    def import_documentos(self, file_path: str, empresa_id: int) -> dict:
        workbook = self._load_workbook(file_path)
        worksheet = workbook.active

        imported = 0
        failed = 0
        errors: list[str] = []

        for row_number, row in self._iter_data_rows(worksheet, self._looks_like_documento_header):
            try:
                if self._is_legacy_document_row(row):
                    meios_recebimento = None
                    nome_documento = self._get_row_value(row, 0)
                    nome_tipo = self._get_row_value(row, 1)
                else:
                    meios_recebimento = self._get_row_value(row, 0)
                    nome_documento = self._get_row_value(row, 1)
                    nome_tipo = self._get_row_value(row, 2)
                self._import_documento_row(empresa_id, meios_recebimento, nome_documento, nome_tipo)
                imported += 1
            except ValidationError as exc:
                failed += 1
                errors.append(f"Linha {row_number}: {exc}")

        return {"imported": imported, "failed": failed, "errors": errors}

    def import_cadastros_completos(self, file_path: str) -> dict:
        workbook = self._load_workbook(file_path)
        worksheet = workbook.active
        import_structure = self._resolve_cadastro_completo_structure(worksheet)

        processed_rows = 0
        documents_imported = 0
        statuses_imported = 0
        failed = 0
        errors: list[str] = []
        company_outcomes: dict[int, str] = {}
        created_type_ids: set[int] = set()
        current_company: dict | None = None
        db_manager = self.empresa_service.empresa_repository.db_manager

        for row_number, row in self._iter_complete_data_rows(worksheet, import_structure["has_header"]):
            next_company: dict | None = None
            try:
                with db_manager.connect():
                    company, outcome = self._resolve_company_for_complete_row(
                        row,
                        current_company,
                        import_structure["field_indexes"],
                    )

                    nome_documento = self._get_mapped_row_value(
                        row,
                        import_structure["field_indexes"],
                        "nome_documento",
                    )
                    nome_tipo = self._get_mapped_row_value(
                        row,
                        import_structure["field_indexes"],
                        "nome_tipo",
                    )
                    meios_recebimento = self._get_mapped_row_value(
                        row,
                        import_structure["field_indexes"],
                        "meios_recebimento",
                    )

                    if self._row_has_values((nome_documento, nome_tipo)):
                        documento_id, tipo_created, tipo_id = self._import_documento_row(
                            company["id"],
                            meios_recebimento,
                            nome_documento,
                            nome_tipo,
                        )
                        if tipo_created:
                            created_type_ids.add(tipo_id)
                        statuses_imported += self._import_status_columns_for_row(
                            documento_id,
                            row,
                            import_structure["status_columns"],
                        )
                        documents_imported += 1

                    self._register_company_outcome(company_outcomes, company["id"], outcome)
                    processed_rows += 1
                    next_company = company
            except ValidationError as exc:
                failed += 1
                errors.append(f"Linha {row_number}: {exc}")
                continue

            current_company = next_company

        companies_created = self._count_company_outcomes(company_outcomes, IMPORT_COMPANY_CREATED)
        companies_updated = self._count_company_outcomes(company_outcomes, IMPORT_COMPANY_UPDATED)
        companies_reused = self._count_company_outcomes(company_outcomes, IMPORT_COMPANY_REUSED)

        return {
            "processed_rows": processed_rows,
            "companies_created": companies_created,
            "companies_updated": companies_updated,
            "companies_reused": companies_reused,
            "documents_imported": documents_imported,
            "statuses_imported": statuses_imported,
            "types_created": len(created_type_ids),
            "failed": failed,
            "errors": errors,
        }

    def _load_workbook(self, file_path: str):
        if load_workbook is None:
            raise ValidationError("A biblioteca openpyxl nao esta instalada no ambiente.")
        try:
            return load_workbook(file_path, read_only=True, data_only=True)
        except Exception as exc:
            raise ValidationError("Nao foi possivel abrir o arquivo Excel informado.") from exc

    def _export_template(
        self,
        file_path: str,
        *,
        worksheet_title: str,
        layout: tuple[dict, ...],
    ) -> dict:
        if Workbook is None:
            raise ValidationError("A biblioteca openpyxl nao esta instalada no ambiente.")

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = worksheet_title

        headers = [item["label"] for item in layout]
        examples = [item["example"] for item in layout]
        worksheet.append(headers)
        worksheet.append(examples)

        if Font is not None:
            for cell in worksheet[1]:
                cell.font = Font(bold=True)

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        for index, item in enumerate(layout, start=1):
            column_letter = worksheet.cell(row=1, column=index).column_letter
            width = max(len(item["label"]), len(str(item["example"] or "")), 18) + 2
            worksheet.column_dimensions[column_letter].width = min(width, 42)

        workbook.save(file_path)
        return {
            "path": file_path,
            "columns": len(layout),
            "example_row": examples,
        }

    def _is_empty_row(self, row) -> bool:
        return all(value is None or str(value).strip() == "" for value in row)

    def _looks_like_empresa_header(self, row) -> bool:
        return self._looks_like_header(row, EMPRESA_IMPORT_HEADER_ALIASES, minimum_matches=2)

    def _looks_like_documento_header(self, row) -> bool:
        return self._looks_like_header(row, DOCUMENTO_IMPORT_HEADER_ALIASES, minimum_matches=2) or self._looks_like_header(
            row,
            LEGACY_DOCUMENTO_IMPORT_HEADER_ALIASES,
            minimum_matches=2,
        )

    def _looks_like_cadastro_completo_header(self, row) -> bool:
        return self._is_complete_header_mapping(self._build_cadastro_completo_header_indexes(row))

    def _looks_like_header(self, row, aliases_by_index: dict[int, set[str]], *, minimum_matches: int) -> bool:
        matches = 0
        for index, aliases in aliases_by_index.items():
            if len(row) <= index:
                continue
            if self._normalize_header_value(row[index]) in aliases:
                matches += 1
        return matches >= minimum_matches

    def _iter_data_rows(self, worksheet, header_checker):
        started_data_block = False
        for row_number, row in enumerate(worksheet.iter_rows(min_row=1, values_only=True), start=1):
            if self._is_empty_row(row):
                continue
            if not started_data_block and header_checker(row):
                started_data_block = True
                continue
            started_data_block = True
            yield row_number, row

    def _iter_complete_data_rows(self, worksheet, has_header: bool):
        skipped_header = False
        for row_number, row in enumerate(worksheet.iter_rows(min_row=1, values_only=True), start=1):
            if self._is_empty_row(row):
                continue
            if has_header and not skipped_header:
                skipped_header = True
                continue
            yield row_number, row

    def _resolve_cadastro_completo_structure(self, worksheet) -> dict:
        for row_number, row in enumerate(worksheet.iter_rows(min_row=1, values_only=True), start=1):
            if self._is_empty_row(row):
                continue

            header_indexes = self._build_cadastro_completo_header_indexes(row)
            if self._is_complete_header_mapping(header_indexes):
                status_columns = self._resolve_status_columns(row, row_number)
                return {
                    "has_header": True,
                    "field_indexes": header_indexes,
                    "status_columns": status_columns,
                }

            return {
                "has_header": False,
                "field_indexes": dict(CADASTRO_COMPLETO_LEGACY_FIELD_INDEXES),
                "status_columns": [],
            }

        return {
            "has_header": False,
            "field_indexes": dict(CADASTRO_COMPLETO_LEGACY_FIELD_INDEXES),
            "status_columns": [],
        }

    def _build_cadastro_completo_header_indexes(self, row) -> dict[str, int]:
        header_indexes: dict[str, int] = {}
        for index, value in enumerate(row):
            normalized_value = self._normalize_header_value(value)
            if not normalized_value:
                continue
            for field_name, aliases in CADASTRO_COMPLETO_FIELD_ALIASES.items():
                if field_name in header_indexes:
                    continue
                if normalized_value in aliases:
                    header_indexes[field_name] = index
                    break
        return header_indexes

    def _is_complete_header_mapping(self, header_indexes: dict[str, int]) -> bool:
        return (
            len(header_indexes) >= 4
            and "codigo_empresa" in header_indexes
            and "nome_empresa" in header_indexes
        )

    def _resolve_status_columns(self, header_row, header_row_number: int) -> list[dict]:
        default_year = self._resolve_default_status_year()
        status_columns: list[dict] = []
        used_periods: dict[tuple[int, int], int] = {}

        for index, raw_value in enumerate(header_row):
            parsed_period = self._parse_status_header(raw_value, default_year)
            if parsed_period is None:
                continue

            year, month = parsed_period
            period_key = (year, month)
            if period_key in used_periods:
                previous_column = used_periods[period_key] + 1
                raise ValidationError(
                    (
                        f'Cabecalho invalido na linha {header_row_number}: '
                        f'o periodo {month:02d}/{year} foi informado mais de uma vez '
                        f"(colunas {previous_column} e {index + 1})."
                    )
                )
            used_periods[period_key] = index
            status_columns.append(
                {
                    "index": index,
                    "year": year,
                    "month": month,
                    "header_label": str(raw_value or MONTH_NAMES[month]).strip() or MONTH_NAMES[month],
                }
            )

        if not status_columns:
            return []

        years = sorted({item["year"] for item in status_columns})
        period_ids = self._ensure_periods_for_import(years)
        for item in status_columns:
            period_id = period_ids.get((item["year"], item["month"]))
            if not period_id:
                raise ValidationError(
                    f'Nao foi possivel localizar o periodo {item["month"]:02d}/{item["year"]} para a importacao.'
                )
            item["periodo_id"] = period_id

        return sorted(status_columns, key=lambda item: (item["year"], item["month"], item["index"]))

    def _resolve_default_status_year(self) -> int:
        available_years = self.periodo_service.list_available_years()
        if len(available_years) == 1:
            return available_years[0]
        return date.today().year

    def _ensure_periods_for_import(self, years: list[int]) -> dict[tuple[int, int], int]:
        for year in years:
            self.periodo_service.generate_year(year)

        period_map: dict[tuple[int, int], int] = {}
        for periodo in self.periodo_service.list_periodos():
            key = (periodo["ano"], periodo["mes"])
            if periodo["ano"] in years:
                period_map[key] = periodo["id"]
        return period_map

    def _parse_status_header(self, value, default_year: int) -> tuple[int, int] | None:
        normalized = self._normalize_header_value(value)
        if not normalized:
            return None

        direct_month = MONTH_HEADER_TO_NUMBER.get(normalized)
        if direct_month is not None:
            return default_year, direct_month

        named_match = re.fullmatch(
            r"(janeiro|fevereiro|marco|abril|maio|junho|julho|agosto|setembro|outubro|novembro|dezembro)_(\d{4})",
            normalized,
        )
        if named_match:
            return int(named_match.group(2)), MONTH_HEADER_TO_NUMBER[named_match.group(1)]

        month_year_match = re.fullmatch(r"(\d{1,2})_(\d{4})", normalized)
        if month_year_match:
            month = int(month_year_match.group(1))
            if 1 <= month <= 12:
                return int(month_year_match.group(2)), month

        year_month_match = re.fullmatch(r"(\d{4})_(\d{1,2})", normalized)
        if year_month_match:
            month = int(year_month_match.group(2))
            if 1 <= month <= 12:
                return int(year_month_match.group(1)), month

        return None

    def _import_empresa_row(self, row) -> int:
        codigo = self._get_row_value(row, 0)
        nome = self._get_row_value(row, 1)
        email = self._get_row_value(row, 2)
        contato = self._get_row_value(row, 3)
        observacao = self._get_row_value(row, 4)
        return self.empresa_service.create_empresa(codigo, nome or "", email, contato, observacao)

    def _import_documento_row(self, empresa_id: int, meios_recebimento, nome_documento, nome_tipo) -> tuple[int, bool, int]:
        if not self._has_value(nome_documento):
            raise ValidationError("O nome do documento deve ser informado.")
        if not self._has_value(nome_tipo):
            raise ValidationError("O tipo do documento deve ser informado.")

        tipo, tipo_created = self.tipo_service.ensure_tipo(str(nome_tipo))
        documento_id = self.documento_service.create_documento(
            empresa_id,
            tipo["id"],
            str(nome_documento or ""),
            meios_recebimento,
        )
        return documento_id, tipo_created, tipo["id"]

    def _import_status_columns_for_row(self, documento_id: int, row, status_columns: list[dict]) -> int:
        imported = 0
        for column in status_columns:
            raw_status = self._get_row_value(row, column["index"])
            normalized_status = self._parse_imported_status(raw_status, column["header_label"])
            if normalized_status is None:
                continue
            self.status_service.update_status(documento_id, column["periodo_id"], normalized_status)
            imported += 1
        return imported

    def _parse_imported_status(self, raw_status, header_label: str) -> str | None:
        if not self._has_value(raw_status):
            return None

        normalized_key = self._normalize_header_value(raw_status)
        status = IMPORT_STATUS_ALIASES.get(normalized_key)
        if status:
            return status

        raise ValidationError(
            (
                f'Valor invalido na coluna "{header_label}". '
                'Use OK para Recebido, P para Pendente, X para Nao cobrar ou deixe em branco.'
            )
        )

    def _resolve_company_for_complete_row(
        self,
        row,
        current_company: dict | None,
        field_indexes: dict[str, int],
    ) -> tuple[dict, str]:
        codigo = self._get_mapped_row_value(row, field_indexes, "codigo_empresa")
        nome = self._get_mapped_row_value(row, field_indexes, "nome_empresa")
        email = self._get_mapped_row_value(row, field_indexes, "email_contato")
        contato = self._get_mapped_row_value(row, field_indexes, "nome_contato")
        observacao = self._get_mapped_row_value(row, field_indexes, "observacao")
        company_values = (codigo, nome, email, contato, observacao)

        if self._has_value(codigo):
            return self._get_or_create_company_for_import(codigo, nome, email, contato, observacao)

        if self._row_has_values(company_values):
            raise ValidationError("Informe o codigo da empresa para identificar o cadastro completo.")

        if current_company is None:
            raise ValidationError(
                "Informe o codigo e o nome da empresa na primeira linha de cada bloco de documentos."
            )

        return current_company, IMPORT_COMPANY_REUSED

    def _get_or_create_company_for_import(
        self,
        codigo,
        nome,
        email,
        contato,
        observacao,
    ) -> tuple[dict, str]:
        existing = self.empresa_service.get_empresa_by_code(codigo, active_only=False)
        if not existing:
            if not self._has_value(nome):
                raise ValidationError("O nome da empresa deve ser informado para novos cadastros.")
            company_id = self.empresa_service.create_empresa(codigo, str(nome or ""), email, contato, observacao)
            return self.empresa_service.get_empresa(company_id), IMPORT_COMPANY_CREATED

        updated_name = self._coalesce_row_value(nome, existing["nome_empresa"])
        updated_email = self._coalesce_row_value(email, existing.get("email_contato"))
        updated_contato = self._coalesce_row_value(contato, existing.get("nome_contato"))
        updated_observacao = self._coalesce_row_value(observacao, existing.get("observacao"))

        if self._company_needs_update(
            existing,
            updated_name,
            updated_email,
            updated_contato,
            updated_observacao,
        ):
            self.empresa_service.update_empresa(
                existing["id"],
                updated_name,
                updated_email,
                updated_contato,
                updated_observacao,
            )
            return self.empresa_service.get_empresa(existing["id"]), IMPORT_COMPANY_UPDATED

        return existing, IMPORT_COMPANY_REUSED

    def _company_needs_update(
        self,
        company: dict,
        nome_empresa: str,
        email_contato: str | None,
        nome_contato: str | None,
        observacao: str | None,
    ) -> bool:
        return any(
            (
                company["nome_empresa"] != nome_empresa,
                (company.get("email_contato") or None) != (email_contato or None),
                (company.get("nome_contato") or None) != (nome_contato or None),
                (company.get("observacao") or None) != (observacao or None),
            )
        )

    def _register_company_outcome(self, outcomes: dict[int, str], company_id: int, status: str) -> None:
        current_status = outcomes.get(company_id)
        if current_status is None:
            outcomes[company_id] = status
            return
        if IMPORT_COMPANY_OUTCOME_PRIORITY[status] > IMPORT_COMPANY_OUTCOME_PRIORITY[current_status]:
            outcomes[company_id] = status

    def _count_company_outcomes(self, outcomes: dict[int, str], status: str) -> int:
        return sum(1 for current_status in outcomes.values() if current_status == status)

    def _get_row_value(self, row, index: int):
        if len(row) <= index:
            return None
        return row[index]

    def _get_mapped_row_value(self, row, field_indexes: dict[str, int], field_name: str):
        index = field_indexes.get(field_name)
        if index is None:
            return None
        return self._get_row_value(row, index)

    def _row_has_values(self, values) -> bool:
        return any(self._has_value(value) for value in values)

    def _has_value(self, value) -> bool:
        return value is not None and str(value).strip() != ""

    def _is_legacy_document_row(self, row) -> bool:
        meaningful_values = [value for value in row if self._has_value(value)]
        return len(meaningful_values) <= 2

    def _coalesce_row_value(self, value, fallback):
        if not self._has_value(value):
            return fallback
        normalized = str(value).strip()
        return normalized or fallback

    def _normalize_header_value(self, value) -> str:
        raw = str(value or "").strip()
        normalized = unicodedata.normalize("NFKD", raw)
        normalized = "".join(char for char in normalized if not unicodedata.combining(char))
        normalized = normalized.casefold()
        normalized = re.sub(r"[^a-z0-9]+", "_", normalized)
        return normalized.strip("_")
