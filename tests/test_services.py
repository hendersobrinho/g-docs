from __future__ import annotations

from datetime import datetime, timedelta, timezone
from pathlib import Path
import tempfile
import unittest
from unittest.mock import patch

from documentos_empresa_app.database.connection import DatabaseManager
from documentos_empresa_app.database.repositories import (
    DeliveryMethodRepository,
    DocumentoRepository,
    EmpresaRepository,
    LogRepository,
    PeriodoRepository,
    RememberedSessionRepository,
    StatusRepository,
    TipoRepository,
    UsuarioRepository,
)
from documentos_empresa_app.database.schema import initialize_schema
from documentos_empresa_app.services.audit_service import AuditService
from documentos_empresa_app.services.auth_service import AuthService
from documentos_empresa_app.services.database_maintenance_service import DatabaseMaintenanceService
from documentos_empresa_app.services.delivery_method_service import DeliveryMethodService
from documentos_empresa_app.services.documento_service import DocumentoService
from documentos_empresa_app.services.empresa_service import EmpresaService
from documentos_empresa_app.services.import_service import ImportService
from documentos_empresa_app.services.log_service import LogService
from documentos_empresa_app.services.pending_report_service import PendingReportService
from documentos_empresa_app.services.periodo_service import PeriodoService
from documentos_empresa_app.services.session_service import SessionService
from documentos_empresa_app.services.status_service import StatusService
from documentos_empresa_app.services.tipo_service import TipoService
from documentos_empresa_app.services.user_service import UserService
from documentos_empresa_app.utils.common import (
    TYPE_OCCURRENCE_ANUAL_JANEIRO,
    TYPE_OCCURRENCE_MENSAL,
    TYPE_OCCURRENCE_TRIMESTRAL,
    ValidationError,
)

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - depende do ambiente local
    load_workbook = None


class ApplicationServiceTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory()
        db_path = Path(self.temp_dir.name) / "test_app.db"

        db_manager = DatabaseManager(db_path)
        initialize_schema(db_manager)

        self.empresa_repository = EmpresaRepository(db_manager)
        self.delivery_method_repository = DeliveryMethodRepository(db_manager)
        self.tipo_repository = TipoRepository(db_manager)
        self.documento_repository = DocumentoRepository(db_manager)
        self.periodo_repository = PeriodoRepository(db_manager)
        self.status_repository = StatusRepository(db_manager)
        self.usuario_repository = UsuarioRepository(db_manager)
        self.remembered_session_repository = RememberedSessionRepository(db_manager)
        self.log_repository = LogRepository(db_manager)

        self.session_service = SessionService()
        self.audit_service = AuditService(self.log_repository, self.session_service)

        self.empresa_service = EmpresaService(
            self.empresa_repository,
            audit_service=self.audit_service,
            session_service=self.session_service,
        )
        self.delivery_method_service = DeliveryMethodService(
            self.delivery_method_repository,
            self.documento_repository,
            audit_service=self.audit_service,
            session_service=self.session_service,
        )
        self.tipo_service = TipoService(self.tipo_repository)
        self.documento_service = DocumentoService(
            self.documento_repository,
            self.empresa_repository,
            self.tipo_repository,
            audit_service=self.audit_service,
            session_service=self.session_service,
        )
        self.periodo_service = PeriodoService(self.periodo_repository)
        self.pending_report_service = PendingReportService(
            self.empresa_repository,
            self.documento_repository,
            self.periodo_repository,
            self.status_repository,
        )
        self.status_service = StatusService(
            self.empresa_repository,
            self.documento_repository,
            self.periodo_repository,
            self.status_repository,
            audit_service=self.audit_service,
            session_service=self.session_service,
        )
        self.import_service = ImportService(
            self.empresa_service,
            self.tipo_service,
            self.documento_service,
            self.periodo_service,
            self.status_service,
        )
        self.database_maintenance_service = DatabaseMaintenanceService(db_manager)
        self.auth_service = AuthService(self.usuario_repository, self.remembered_session_repository)
        self.user_service = UserService(
            self.usuario_repository,
            self.session_service,
            audit_service=self.audit_service,
            auth_service=self.auth_service,
        )
        self.log_service = LogService(self.log_repository, self.session_service)

        admin_user = self.auth_service.authenticate("admin", "admin")
        self.session_service.login(admin_user)

    def tearDown(self) -> None:
        self.temp_dir.cleanup()

    def test_initial_types_are_seeded(self) -> None:
        tipos = [item["nome_tipo"] for item in self.tipo_service.list_tipos()]
        self.assertIn("Extratos CC", tipos)
        self.assertIn("Contratos", tipos)

    def test_initial_delivery_methods_are_seeded(self) -> None:
        methods = [item["nome_meio"] for item in self.delivery_method_service.list_methods()]
        self.assertIn("WhatsApp", methods)
        self.assertIn("Email", methods)

    def test_generate_year_does_not_duplicate_months(self) -> None:
        first_run = self.periodo_service.generate_year(2026)
        second_run = self.periodo_service.generate_year(2026)

        self.assertEqual(first_run["created"], 12)
        self.assertEqual(second_run["created"], 0)
        self.assertEqual(second_run["existing"], 12)

    def test_closure_hides_future_periods(self) -> None:
        empresa_id = self.empresa_service.create_empresa(101, "Empresa Teste")
        tipo_id = self.tipo_service.get_or_create_tipo("Documentos Financeiros")["id"]
        documento_id = self.documento_service.create_documento(empresa_id, tipo_id, "Banco X")
        self.periodo_service.generate_year(2026)

        periodos = self.periodo_service.list_periodos()
        jan = next(item for item in periodos if item["ano"] == 2026 and item["mes"] == 1)
        apr = next(item for item in periodos if item["ano"] == 2026 and item["mes"] == 4)
        jun = next(item for item in periodos if item["ano"] == 2026 and item["mes"] == 6)

        self.status_service.update_status(documento_id, jun["id"], "Recebido")
        self.status_service.update_status(documento_id, apr["id"], "Encerrado")

        view = self.status_service.build_control_view(empresa_id, jan["id"], jun["id"])
        documentos = view["groups"][0]["documentos"]
        availability = [cell["available"] for cell in documentos[0]["cells"]]

        self.assertEqual(availability, [True, True, True, True, False, False])
        self.assertEqual(documentos[0]["cells"][3]["status"], "Encerrado")
        self.assertEqual(documentos[0]["cells"][5]["status"], "")

        may = next(item for item in periodos if item["ano"] == 2026 and item["mes"] == 5)
        later_view = self.status_service.build_control_view(empresa_id, may["id"], jun["id"])
        self.assertEqual(later_view["groups"], [])

    def test_pending_report_lists_only_pending_statuses(self) -> None:
        empresa_id = self.empresa_service.create_empresa(110, "Empresa Pendencias")
        tipo_id = self.tipo_service.get_or_create_tipo("Extratos CC")["id"]
        documento_a_id = self.documento_service.create_documento(empresa_id, tipo_id, "Banco A")
        documento_b_id = self.documento_service.create_documento(empresa_id, tipo_id, "Banco B")
        self.periodo_service.generate_year(2026)

        periodos = self.periodo_service.list_periodos()
        jan = next(item for item in periodos if item["ano"] == 2026 and item["mes"] == 1)
        feb = next(item for item in periodos if item["ano"] == 2026 and item["mes"] == 2)

        self.status_service.update_status(documento_a_id, jan["id"], "Pendente")
        self.status_service.update_status(documento_a_id, feb["id"], "Recebido")
        self.status_service.update_status(documento_b_id, feb["id"], "Pendente")

        report = self.pending_report_service.list_pending_rows([empresa_id], jan["id"], feb["id"])

        self.assertEqual(len(report["rows"]), 2)
        self.assertEqual(
            [(row["nome_documento"], row["mes"], row["status"]) for row in report["rows"]],
            [("Banco A", 1, "Pendente"), ("Banco B", 2, "Pendente")],
        )

    def test_status_view_includes_last_change_metadata(self) -> None:
        empresa_id = self.empresa_service.create_empresa(112, "Empresa Status")
        tipo_id = self.tipo_service.get_or_create_tipo("Extratos CC")["id"]
        documento_id = self.documento_service.create_documento(empresa_id, tipo_id, "Banco Auditoria")
        self.periodo_service.generate_year(2026)

        periodos = self.periodo_service.list_periodos()
        jan = next(item for item in periodos if item["ano"] == 2026 and item["mes"] == 1)

        self.status_service.update_status(documento_id, jan["id"], "Recebido")

        view = self.status_service.build_control_view(empresa_id, jan["id"], jan["id"])
        cell = view["groups"][0]["documentos"][0]["cells"][0]

        self.assertEqual(cell["status"], "Recebido")
        self.assertEqual(cell["updated_by_username"], "admin")
        self.assertTrue(cell["updated_at"])

    def test_batch_status_update_applies_same_status_to_multiple_documents(self) -> None:
        empresa_id = self.empresa_service.create_empresa(113, "Empresa Lote")
        tipo_id = self.tipo_service.get_or_create_tipo("Extratos CC")["id"]
        documento_a_id = self.documento_service.create_documento(empresa_id, tipo_id, "Banco Lote A")
        documento_b_id = self.documento_service.create_documento(empresa_id, tipo_id, "Banco Lote B")
        self.periodo_service.generate_year(2026)

        periodos = self.periodo_service.list_periodos()
        jan = next(item for item in periodos if item["ano"] == 2026 and item["mes"] == 1)

        result = self.status_service.update_status_batch([documento_a_id, documento_b_id], jan["id"], "Recebido")
        view = self.status_service.build_control_view(empresa_id, jan["id"], jan["id"])
        documentos = {
            documento["nome_documento"]: documento
            for group in view["groups"]
            for documento in group["documentos"]
        }

        self.assertEqual(result["selected"], 2)
        self.assertEqual(result["updated"], 2)
        self.assertEqual(documentos["Banco Lote A"]["cells"][0]["status"], "Recebido")
        self.assertEqual(documentos["Banco Lote B"]["cells"][0]["status"], "Recebido")

    def test_batch_status_update_rolls_back_when_one_document_is_invalid(self) -> None:
        empresa_id = self.empresa_service.create_empresa(114, "Empresa Lote Invalido")
        tipo_mensal_id = self.tipo_service.get_or_create_tipo("Extratos CC")["id"]
        tipo_anual_id = self.tipo_service.create_tipo("Declaracao Anual", TYPE_OCCURRENCE_ANUAL_JANEIRO)
        documento_mensal_id = self.documento_service.create_documento(empresa_id, tipo_mensal_id, "Banco Mensal")
        documento_anual_id = self.documento_service.create_documento(empresa_id, tipo_anual_id, "Declaracao")
        self.periodo_service.generate_year(2026)

        periodos = self.periodo_service.list_periodos()
        fevereiro = next(item for item in periodos if item["ano"] == 2026 and item["mes"] == 2)

        with self.assertRaises(ValidationError):
            self.status_service.update_status_batch(
                [documento_mensal_id, documento_anual_id],
                fevereiro["id"],
                "Recebido",
            )

        self.assertIsNone(self.status_repository.get_by_document_and_period(documento_mensal_id, fevereiro["id"]))
        self.assertIsNone(self.status_repository.get_by_document_and_period(documento_anual_id, fevereiro["id"]))

    @unittest.skipIf(load_workbook is None, "openpyxl nao esta disponivel no ambiente de teste")
    def test_pending_report_exports_excel(self) -> None:
        empresa_id = self.empresa_service.create_empresa(111, "Empresa Excel")
        tipo_id = self.tipo_service.get_or_create_tipo("Contratos")["id"]
        documento_id = self.documento_service.create_documento(empresa_id, tipo_id, "Contrato Social")
        self.periodo_service.generate_year(2026)
        periodos = self.periodo_service.list_periodos()
        mar = next(item for item in periodos if item["ano"] == 2026 and item["mes"] == 3)

        self.status_service.update_status(documento_id, mar["id"], "Pendente")

        file_path = Path(self.temp_dir.name) / "relatorio_pendencias.xlsx"
        result = self.pending_report_service.export_pending_report(str(file_path), [empresa_id], mar["id"], mar["id"])

        workbook = load_workbook(file_path)
        worksheet = workbook["Pendencias"]

        self.assertEqual(result["rows"], 1)
        self.assertEqual(worksheet.max_row, 2)
        self.assertEqual(worksheet.max_column, 4)
        self.assertEqual(worksheet["A1"].value, "Empresa")
        self.assertEqual(worksheet["B1"].value, "Periodo")
        self.assertEqual(worksheet["C1"].value, "Documento pendente")
        self.assertEqual(worksheet["D1"].value, "Status")
        self.assertEqual(worksheet["A2"].value, "Empresa Excel")
        self.assertEqual(worksheet["B2"].value, "03/2026 - Marco")
        self.assertEqual(worksheet["C2"].value, "Contrato Social")
        self.assertEqual(worksheet["D2"].value, "Pendente")

    def test_backup_restore_recovers_previous_database_state(self) -> None:
        self.empresa_service.create_empresa(113, "Empresa Antes do Backup")
        backup_path = Path(self.temp_dir.name) / "backup_teste.db"

        self.database_maintenance_service.create_backup(backup_path)
        self.empresa_service.create_empresa(114, "Empresa Depois do Backup")

        self.database_maintenance_service.restore_backup(backup_path)
        empresas = self.empresa_service.list_empresas(active_only=False)
        codigos = [empresa["codigo_empresa"] for empresa in empresas]

        self.assertIn(113, codigos)
        self.assertNotIn(114, codigos)

    def test_database_manager_applies_sqlite_pragmas(self) -> None:
        with self.empresa_repository.db_manager.connect() as connection:
            foreign_keys = int(connection.execute("PRAGMA foreign_keys").fetchone()[0])
            busy_timeout = int(connection.execute("PRAGMA busy_timeout").fetchone()[0])
            journal_mode = str(connection.execute("PRAGMA journal_mode").fetchone()[0]).lower()
            synchronous = int(connection.execute("PRAGMA synchronous").fetchone()[0])

        self.assertEqual(foreign_keys, 1)
        self.assertEqual(busy_timeout, 5000)
        self.assertEqual(journal_mode, "wal")
        self.assertEqual(synchronous, 1)

    def test_database_maintenance_service_optimize_database_returns_summary(self) -> None:
        self.empresa_service.create_empresa(401, "Empresa Otimizada")

        result = self.database_maintenance_service.optimize_database()

        self.assertTrue(result["optimized"])
        self.assertGreater(result["page_count"], 0)
        self.assertGreaterEqual(result["free_pages"], 0)

    def test_repository_batch_queries_handle_large_repeated_id_lists(self) -> None:
        empresa_id = self.empresa_service.create_empresa(402, "Empresa Lote")
        tipo_id = self.tipo_service.get_or_create_tipo("Extratos CC")["id"]
        documento_id = self.documento_service.create_documento(empresa_id, tipo_id, "Banco Lote")
        self.periodo_service.generate_year(2026)
        janeiro = next(
            periodo
            for periodo in self.periodo_service.list_periodos()
            if periodo["ano"] == 2026 and periodo["mes"] == 1
        )
        self.status_service.update_status(documento_id, janeiro["id"], "Pendente")

        empresas = self.empresa_repository.list_by_ids([empresa_id] * 1200)
        documentos = self.documento_repository.list_by_company_ids([empresa_id] * 1200)
        statuses = self.status_repository.list_for_documents_and_periods([documento_id] * 1200, [janeiro["id"]])

        self.assertEqual(len(empresas), 1)
        self.assertEqual(len(documentos), 1)
        self.assertEqual(len(statuses), 1)
        self.assertEqual(statuses[0]["status"], "Pendente")

    @unittest.skipIf(load_workbook is None, "openpyxl nao esta disponivel no ambiente de teste")
    def test_import_service_exports_empresa_template(self) -> None:
        file_path = Path(self.temp_dir.name) / "modelo_empresas.xlsx"

        result = self.import_service.export_empresa_template(str(file_path))
        workbook = load_workbook(file_path)
        worksheet = workbook["Empresas"]

        self.assertEqual(result["columns"], 5)
        self.assertEqual(worksheet["A1"].value, "Codigo da empresa")
        self.assertEqual(worksheet["E1"].value, "Observacao")
        self.assertEqual(worksheet["E2"].value, "Prefere retorno ate o dia 10")

    @unittest.skipIf(load_workbook is None, "openpyxl nao esta disponivel no ambiente de teste")
    def test_import_service_exports_document_template(self) -> None:
        file_path = Path(self.temp_dir.name) / "modelo_documentos.xlsx"

        result = self.import_service.export_documento_template(str(file_path))
        workbook = load_workbook(file_path)
        worksheet = workbook["Documentos"]

        self.assertEqual(result["columns"], 3)
        self.assertEqual(worksheet["A1"].value, "Meios de recebimento")
        self.assertEqual(worksheet["C2"].value, "Extratos CC")

    @unittest.skipIf(load_workbook is None, "openpyxl nao esta disponivel no ambiente de teste")
    def test_import_service_exports_complete_registration_template(self) -> None:
        file_path = Path(self.temp_dir.name) / "modelo_cadastro_completo.xlsx"

        result = self.import_service.export_cadastro_completo_template(str(file_path))
        workbook = load_workbook(file_path)
        worksheet = workbook["Cadastros completos"]

        self.assertEqual(result["columns"], 20)
        self.assertEqual(worksheet["A1"].value, "Codigo da empresa")
        self.assertEqual(worksheet["E1"].value, "Nome do documento")
        self.assertEqual(worksheet["F1"].value, "Meio de recebimento")
        self.assertEqual(worksheet["H1"].value, "Observacao")
        self.assertEqual(worksheet["I1"].value, "Janeiro")
        self.assertEqual(worksheet["J1"].value, "Fevereiro")
        self.assertEqual(worksheet["K1"].value, "Marco")
        self.assertEqual(worksheet["H2"].value, "Prefere retorno ate o dia 10")
        self.assertEqual(worksheet["I2"].value, "OK")
        self.assertEqual(worksheet["J2"].value, "P")
        self.assertEqual(worksheet["K2"].value, "X")

    def test_empresa_service_accepts_integer_like_numeric_codes(self) -> None:
        empresa_id = self.empresa_service.create_empresa("101.0", "Empresa Decimal")
        empresa = self.empresa_service.get_empresa(empresa_id)

        self.assertEqual(empresa["codigo_empresa"], 101)

    def test_empresa_service_saves_contact_fields(self) -> None:
        empresa_id = self.empresa_service.create_empresa(
            102,
            "Empresa Campos",
            email_contato="contato@empresa.com",
            nome_contato="Maria Silva",
        )

        empresa = self.empresa_service.get_empresa(empresa_id)

        self.assertIsNone(empresa["meios_recebimento"])
        self.assertEqual(empresa["email_contato"], "contato@empresa.com")
        self.assertEqual(empresa["nome_contato"], "Maria Silva")

    def test_empresa_service_saves_observacao_and_limits_length(self) -> None:
        empresa_id = self.empresa_service.create_empresa(
            115,
            "Empresa Observacao",
            observacao="Cliente prefere consolidado mensal.",
        )
        empresa = self.empresa_service.get_empresa(empresa_id)

        self.assertEqual(empresa["observacao"], "Cliente prefere consolidado mensal.")

        with self.assertRaises(ValidationError):
            self.empresa_service.update_empresa(
                empresa_id,
                "Empresa Observacao",
                observacao="x" * 256,
            )

    def test_update_empresa_nome_preserves_observacao(self) -> None:
        empresa_id = self.empresa_service.create_empresa(
            116,
            "Empresa Original",
            observacao="Nao perder esta observacao.",
        )

        self.empresa_service.update_empresa_nome(empresa_id, "Empresa Renomeada")
        empresa = self.empresa_service.get_empresa(empresa_id)

        self.assertEqual(empresa["nome_empresa"], "Empresa Renomeada")
        self.assertEqual(empresa["observacao"], "Nao perder esta observacao.")

    def test_documento_service_preserves_delivery_methods(self) -> None:
        empresa_id = self.empresa_service.create_empresa(
            104,
            "Empresa Meio Legado",
        )
        tipo_id = self.tipo_service.get_or_create_tipo("Comprovantes")["id"]
        documento_id = self.documento_service.create_documento(
            empresa_id,
            tipo_id,
            "Portal Fiscal",
            ["Portal", "email", "portal"],
        )

        documento = self.documento_service.get_documento(documento_id)

        self.assertEqual(documento["meios_recebimento"], "Portal, Email")

    def test_document_name_suggestions_only_use_system_standard_names(self) -> None:
        empresa_id = self.empresa_service.create_empresa(122, "Empresa Sugestoes")
        tipo_extrato_id = self.tipo_service.get_or_create_tipo("Extratos CC")["id"]
        tipo_contrato_id = self.tipo_service.get_or_create_tipo("Contratos")["id"]

        self.documento_service.create_documento(empresa_id, tipo_extrato_id, "Banco Importado")
        self.assertEqual(self.documento_service.list_document_name_suggestions(tipo_extrato_id), [])

        self.documento_service.create_system_document_name(tipo_extrato_id, "Banco Padrao")
        self.documento_service.create_system_document_name(tipo_contrato_id, "Contrato Social")

        self.assertEqual(
            self.documento_service.list_document_name_suggestions(tipo_extrato_id),
            ["Banco Padrao"],
        )
        self.assertEqual(
            self.documento_service.list_document_name_suggestions(search="Contrato"),
            ["Contrato Social"],
        )

    def test_custom_delivery_methods_do_not_enter_system_method_list(self) -> None:
        empresa_id = self.empresa_service.create_empresa(123, "Empresa Meio Livre")
        tipo_id = self.tipo_service.get_or_create_tipo("Comprovantes")["id"]

        self.documento_service.create_documento(
            empresa_id,
            tipo_id,
            "Portal Fiscal",
            ["Portal do cliente"],
        )

        methods = [item["nome_meio"] for item in self.delivery_method_service.list_methods()]
        self.assertNotIn("Portal do cliente", methods)

    def test_schema_migrates_legacy_company_delivery_methods_to_documents(self) -> None:
        empresa_id = self.empresa_repository.create(117, "Empresa Legada", "Email, Onvio")
        tipo_id = self.tipo_service.get_or_create_tipo("Extratos CC")["id"]
        documento_id = self.documento_repository.create(empresa_id, tipo_id, None, "Banco Legado")

        initialize_schema(self.empresa_repository.db_manager)

        documento = self.documento_service.get_documento(documento_id)

        self.assertEqual(documento["meios_recebimento"], "Email, Onvio")

    def test_schema_migrates_legacy_status_table_to_allow_nao_cobrar(self) -> None:
        legacy_db_path = Path(self.temp_dir.name) / "legacy_status.db"
        legacy_db_manager = DatabaseManager(legacy_db_path)

        with legacy_db_manager.connect() as connection:
            connection.execute(
                """
                CREATE TABLE status_documento_mensal (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    documento_empresa_id INTEGER NOT NULL,
                    periodo_id INTEGER NOT NULL,
                    status TEXT NULL CHECK (status IN ('Recebido', 'Pendente', 'Encerrado') OR status IS NULL),
                    FOREIGN KEY (documento_empresa_id) REFERENCES documentos_empresa(id) ON DELETE CASCADE,
                    FOREIGN KEY (periodo_id) REFERENCES periodos(id) ON DELETE CASCADE,
                    CONSTRAINT uq_status UNIQUE (documento_empresa_id, periodo_id)
                )
                """
            )

        initialize_schema(legacy_db_manager)

        with legacy_db_manager.connect() as connection:
            table_sql = connection.execute(
                """
                SELECT sql
                FROM sqlite_master
                WHERE type = 'table' AND name = 'status_documento_mensal'
                """
            ).fetchone()["sql"]

        self.assertIn("'Nao cobrar'", table_sql)

    def test_empresa_service_can_save_company_directory(self) -> None:
        empresa_id = self.empresa_service.create_empresa(108, "Empresa Pasta")

        self.empresa_service.set_empresa_directory(empresa_id, "~/Documentos/Empresa Pasta")
        empresa = self.empresa_service.get_empresa(empresa_id)

        self.assertTrue(empresa["diretorio_documentos"])
        self.assertIn("Empresa Pasta", empresa["diretorio_documentos"])

    def test_empresa_update_preserves_saved_directory(self) -> None:
        empresa_id = self.empresa_service.create_empresa(109, "Empresa Diretorio")
        self.empresa_service.set_empresa_directory(empresa_id, "~/Documentos/Empresa Diretorio")

        self.empresa_service.update_empresa(
            empresa_id,
            "Empresa Diretorio Atualizada",
            "diretorio@empresa.com",
            "Rita",
        )
        empresa = self.empresa_service.get_empresa(empresa_id)

        self.assertEqual(empresa["nome_empresa"], "Empresa Diretorio Atualizada")
        self.assertTrue(empresa["diretorio_documentos"])
        self.assertIn("Empresa Diretorio", empresa["diretorio_documentos"])

    def test_empresa_service_rejects_invalid_optional_email(self) -> None:
        with self.assertRaises(ValidationError):
            self.empresa_service.create_empresa(
                103,
                "Empresa Email Invalido",
                email_contato="email_invalido",
            )

    def test_tipo_service_blocks_delete_when_type_is_in_use(self) -> None:
        empresa_id = self.empresa_service.create_empresa(105, "Empresa Tipo Em Uso")
        tipo_id = self.tipo_service.get_or_create_tipo("Comprovantes")["id"]
        self.documento_service.create_documento(empresa_id, tipo_id, "Portal Fiscal")

        with self.assertRaises(ValidationError):
            self.tipo_service.delete_tipo(tipo_id)

    def test_tipo_service_saves_occurrence_rule(self) -> None:
        tipo_id = self.tipo_service.create_tipo("Informes", TYPE_OCCURRENCE_TRIMESTRAL)
        tipo = self.tipo_service.get_tipo(tipo_id)

        self.assertEqual(tipo["regra_ocorrencia"], TYPE_OCCURRENCE_TRIMESTRAL)

        self.tipo_service.update_tipo(tipo_id, "Informes anuais", TYPE_OCCURRENCE_ANUAL_JANEIRO)
        tipo_atualizado = self.tipo_service.get_tipo(tipo_id)

        self.assertEqual(tipo_atualizado["nome_tipo"], "Informes anuais")
        self.assertEqual(tipo_atualizado["regra_ocorrencia"], TYPE_OCCURRENCE_ANUAL_JANEIRO)

    def test_quarterly_type_marks_non_chargeable_months_as_nao_cobrar(self) -> None:
        empresa_id = self.empresa_service.create_empresa(118, "Empresa Trimestral")
        tipo_id = self.tipo_service.create_tipo("Informe trimestral", TYPE_OCCURRENCE_TRIMESTRAL)
        documento_id = self.documento_service.create_documento(empresa_id, tipo_id, "Informe IR")
        self.periodo_service.generate_year(2026)

        periodos = self.periodo_service.list_periodos()
        jan = next(item for item in periodos if item["ano"] == 2026 and item["mes"] == 1)
        apr = next(item for item in periodos if item["ano"] == 2026 and item["mes"] == 4)
        feb = next(item for item in periodos if item["ano"] == 2026 and item["mes"] == 2)

        self.status_repository.upsert(documento_id, feb["id"], "Encerrado", self.session_service.get_user_id())
        self.status_service.update_status(documento_id, apr["id"], "Pendente")

        view = self.status_service.build_control_view(empresa_id, jan["id"], apr["id"])
        cells = view["groups"][0]["documentos"][0]["cells"]

        self.assertEqual([cell["available"] for cell in cells], [True, False, False, True])
        self.assertEqual([cell["status"] for cell in cells], ["", "Nao cobrar", "Nao cobrar", "Pendente"])

    def test_quarterly_type_rejects_status_update_outside_chargeable_months(self) -> None:
        empresa_id = self.empresa_service.create_empresa(119, "Empresa Bloqueio Trimestral")
        tipo_id = self.tipo_service.create_tipo("Documento trimestral", TYPE_OCCURRENCE_TRIMESTRAL)
        documento_id = self.documento_service.create_documento(empresa_id, tipo_id, "Apuracao")
        self.periodo_service.generate_year(2026)

        fevereiro = next(
            item for item in self.periodo_service.list_periodos() if item["ano"] == 2026 and item["mes"] == 2
        )

        with self.assertRaises(ValidationError):
            self.status_service.update_status(documento_id, fevereiro["id"], "Recebido")

    def test_annual_january_type_only_enables_january_cell(self) -> None:
        empresa_id = self.empresa_service.create_empresa(120, "Empresa Anual")
        tipo_id = self.tipo_service.create_tipo("Certidao anual", TYPE_OCCURRENCE_ANUAL_JANEIRO)
        documento_id = self.documento_service.create_documento(empresa_id, tipo_id, "Certidao negativa")
        self.periodo_service.generate_year(2026)

        periodos = self.periodo_service.list_periodos()
        jan = next(item for item in periodos if item["ano"] == 2026 and item["mes"] == 1)
        mar = next(item for item in periodos if item["ano"] == 2026 and item["mes"] == 3)
        self.status_service.update_status(documento_id, jan["id"], "Recebido")

        view = self.status_service.build_control_view(empresa_id, jan["id"], mar["id"])
        cells = view["groups"][0]["documentos"][0]["cells"]

        self.assertEqual([cell["available"] for cell in cells], [True, False, False])
        self.assertEqual([cell["status"] for cell in cells], ["Recebido", "Nao cobrar", "Nao cobrar"])

    def test_pending_report_ignores_pending_status_in_non_chargeable_month(self) -> None:
        empresa_id = self.empresa_service.create_empresa(121, "Empresa Relatorio Especial")
        tipo_id = self.tipo_service.create_tipo("Documento especial", TYPE_OCCURRENCE_ANUAL_JANEIRO)
        documento_id = self.documento_service.create_documento(empresa_id, tipo_id, "Declaracao anual")
        self.periodo_service.generate_year(2026)

        periodos = self.periodo_service.list_periodos()
        janeiro = next(item for item in periodos if item["ano"] == 2026 and item["mes"] == 1)
        fevereiro = next(item for item in periodos if item["ano"] == 2026 and item["mes"] == 2)

        self.status_repository.upsert(documento_id, fevereiro["id"], "Pendente", self.session_service.get_user_id())
        self.status_service.update_status(documento_id, janeiro["id"], "Pendente")

        report = self.pending_report_service.list_pending_rows([empresa_id], janeiro["id"], fevereiro["id"])

        self.assertEqual(
            [(row["nome_documento"], row["mes"]) for row in report["rows"]],
            [("Declaracao anual", 1)],
        )

    def test_renaming_delivery_method_updates_documents_using_it(self) -> None:
        empresa_id = self.empresa_service.create_empresa(
            106,
            "Empresa Meio Global",
        )
        tipo_id = self.tipo_service.get_or_create_tipo("Extratos CC")["id"]
        documento_id = self.documento_service.create_documento(
            empresa_id,
            tipo_id,
            "Banco Principal",
            ["Email", "Onvio"],
        )
        method = next(item for item in self.delivery_method_service.list_methods() if item["nome_meio"] == "Email")

        affected = self.delivery_method_service.update_method(method["id"], "E-mail")
        documento = self.documento_service.get_documento(documento_id)
        methods = [item["nome_meio"] for item in self.delivery_method_service.list_methods()]

        self.assertEqual(affected, 1)
        self.assertEqual(documento["meios_recebimento"], "E-mail, Onvio")
        self.assertIn("E-mail", methods)
        self.assertNotIn("Email", methods)

    def test_deleting_delivery_method_only_removes_it_from_system_list(self) -> None:
        empresa_id = self.empresa_service.create_empresa(
            107,
            "Empresa Meio Removido",
        )
        tipo_id = self.tipo_service.get_or_create_tipo("Comprovantes")["id"]
        documento_id = self.documento_service.create_documento(
            empresa_id,
            tipo_id,
            "Comprovante Portal",
            ["Onvio"],
        )
        method = next(item for item in self.delivery_method_service.list_methods() if item["nome_meio"] == "Onvio")

        affected = self.delivery_method_service.delete_method(method["id"])
        documento = self.documento_service.get_documento(documento_id)
        methods = [item["nome_meio"] for item in self.delivery_method_service.list_methods()]

        self.assertEqual(affected, 1)
        self.assertEqual(documento["meios_recebimento"], "Onvio")
        self.assertNotIn("Onvio", methods)

    def test_sparse_period_range_still_blocks_more_than_twelve_months(self) -> None:
        start_id = self.periodo_repository.create(2026, 1)
        end_id = self.periodo_repository.create(2027, 12)

        with self.assertRaises(ValidationError):
            self.periodo_service.get_periods_between(start_id, end_id)

    def test_import_service_wraps_excel_open_errors(self) -> None:
        with patch("documentos_empresa_app.services.import_service.load_workbook", side_effect=OSError("boom")):
            with self.assertRaises(ValidationError):
                self.import_service.import_empresas("arquivo_invalido.xlsx")

    def test_import_service_imports_company_optional_fields_and_skips_header(self) -> None:
        class FakeWorksheet:
            def __init__(self, rows):
                self._rows = rows

            def iter_rows(self, min_row=1, values_only=True):
                return iter(self._rows)

        class FakeWorkbook:
            def __init__(self, rows):
                self.active = FakeWorksheet(rows)

        workbook = FakeWorkbook(
            [
                (
                    "codigo_empresa",
                    "nome_empresa",
                    "email_contato",
                    "nome_contato",
                    "observacao",
                ),
                (201, "Empresa Importada", "importacao@empresa.com", "Joana", "Observacao importada"),
            ]
        )

        with patch.object(self.import_service, "_load_workbook", return_value=workbook):
            result = self.import_service.import_empresas("empresas.xlsx")

        empresa = self.empresa_service.get_empresa_by_code(201)

        self.assertEqual(result["imported"], 1)
        self.assertEqual(result["failed"], 0)
        self.assertEqual(empresa["nome_empresa"], "Empresa Importada")
        self.assertIsNone(empresa["meios_recebimento"])
        self.assertEqual(empresa["email_contato"], "importacao@empresa.com")
        self.assertEqual(empresa["nome_contato"], "Joana")
        self.assertEqual(empresa["observacao"], "Observacao importada")

    def test_import_service_keeps_legacy_company_two_column_layout(self) -> None:
        class FakeWorksheet:
            def __init__(self, rows):
                self._rows = rows

            def iter_rows(self, min_row=1, values_only=True):
                return iter(self._rows)

        class FakeWorkbook:
            def __init__(self, rows):
                self.active = FakeWorksheet(rows)

        workbook = FakeWorkbook(
            [
                (301, "Empresa Antiga"),
            ]
        )

        with patch.object(self.import_service, "_load_workbook", return_value=workbook):
            result = self.import_service.import_empresas("empresas_antigas.xlsx")

        empresa = self.empresa_service.get_empresa_by_code(301)

        self.assertEqual(result["imported"], 1)
        self.assertEqual(result["failed"], 0)
        self.assertEqual(empresa["nome_empresa"], "Empresa Antiga")
        self.assertIsNone(empresa["meios_recebimento"])
        self.assertIsNone(empresa["email_contato"])
        self.assertIsNone(empresa["nome_contato"])

    def test_import_service_imports_documents_and_skips_header(self) -> None:
        class FakeWorksheet:
            def __init__(self, rows):
                self._rows = rows

            def iter_rows(self, min_row=1, values_only=True):
                return iter(self._rows)

        class FakeWorkbook:
            def __init__(self, rows):
                self.active = FakeWorksheet(rows)

        empresa_id = self.empresa_service.create_empresa(401, "Empresa Docs")
        workbook = FakeWorkbook(
            [
                ("meios_recebimento", "nome_documento", "nome_tipo"),
                ("Email, Portal", "Banco XP", "Extratos CC"),
            ]
        )

        with patch.object(self.import_service, "_load_workbook", return_value=workbook):
            result = self.import_service.import_documentos("documentos.xlsx", empresa_id)

        documentos = self.documento_service.list_documentos_empresa(empresa_id)

        self.assertEqual(result["imported"], 1)
        self.assertEqual(result["failed"], 0)
        self.assertEqual(len(documentos), 1)
        self.assertEqual(documentos[0]["nome_documento"], "Banco XP")
        self.assertEqual(documentos[0]["meios_recebimento"], "Email, Portal")
        self.assertEqual(documentos[0]["nome_tipo"], "Extratos CC")

    def test_import_service_keeps_legacy_document_two_column_layout(self) -> None:
        class FakeWorksheet:
            def __init__(self, rows):
                self._rows = rows

            def iter_rows(self, min_row=1, values_only=True):
                return iter(self._rows)

        class FakeWorkbook:
            def __init__(self, rows):
                self.active = FakeWorksheet(rows)

        empresa_id = self.empresa_service.create_empresa(402, "Empresa Docs Legado")
        workbook = FakeWorkbook(
            [
                ("nome_documento", "nome_tipo"),
                ("Banco Antigo", "Extratos CC"),
            ]
        )

        with patch.object(self.import_service, "_load_workbook", return_value=workbook):
            result = self.import_service.import_documentos("documentos_legados.xlsx", empresa_id)

        documentos = self.documento_service.list_documentos_empresa(empresa_id)

        self.assertEqual(result["imported"], 1)
        self.assertEqual(result["failed"], 0)
        self.assertEqual(len(documentos), 1)
        self.assertEqual(documentos[0]["nome_documento"], "Banco Antigo")
        self.assertIsNone(documentos[0]["meios_recebimento"])
        self.assertEqual(documentos[0]["nome_tipo"], "Extratos CC")

    def test_document_service_lists_reusable_document_names_by_type(self) -> None:
        tipo_extrato_id = self.tipo_service.get_or_create_tipo("Extratos CC")["id"]
        tipo_contrato_id = self.tipo_service.get_or_create_tipo("Contratos")["id"]

        self.documento_service.create_system_document_name(tipo_extrato_id, "Banco do Brasil")
        self.documento_service.create_system_document_name(tipo_extrato_id, "Bradesco")
        self.documento_service.create_system_document_name(tipo_contrato_id, "Contrato Social")

        extrato_names = self.documento_service.list_document_name_suggestions(tipo_extrato_id)
        contrato_names = self.documento_service.list_document_name_suggestions(tipo_contrato_id)
        filtered_names = self.documento_service.list_document_name_suggestions(tipo_extrato_id, search="brad")

        self.assertEqual(extrato_names, ["Banco do Brasil", "Bradesco"])
        self.assertEqual(contrato_names, ["Contrato Social"])
        self.assertEqual(filtered_names, ["Bradesco"])

    def test_import_service_imports_complete_registrations_for_multiple_companies(self) -> None:
        class FakeWorksheet:
            def __init__(self, rows):
                self._rows = rows

            def iter_rows(self, min_row=1, values_only=True):
                return iter(self._rows)

        class FakeWorkbook:
            def __init__(self, rows):
                self.active = FakeWorksheet(rows)

        self.periodo_service.generate_year(2026)
        workbook = FakeWorkbook(
            [
                (
                    "codigo_empresa",
                    "nome_empresa",
                    "email_contato",
                    "nome_contato",
                    "nome_documento",
                    "meio de recebimento",
                    "tipos",
                    "observacoes",
                    "janeiro",
                    "fevereiro",
                    "marco",
                ),
                (
                    601,
                    "Empresa Importacao Completa A",
                    "empresa.a@teste.com",
                    "Ana",
                    "Banco A",
                    "Email",
                    "Extratos CC",
                    "Observacao A",
                    "OK",
                    "P",
                    "X",
                ),
                (None, None, None, None, "Relatorio Gerencial", "Portal", "Balancetes", None, None, None, None),
                (
                    602,
                    "Empresa Importacao Completa B",
                    "empresa.b@teste.com",
                    "Bruno",
                    None,
                    None,
                    None,
                    "Observacao B",
                    None,
                    None,
                    None,
                ),
            ]
        )

        with patch.object(self.import_service, "_load_workbook", return_value=workbook):
            result = self.import_service.import_cadastros_completos("cadastro_completo.xlsx")

        empresa_a = self.empresa_service.get_empresa_by_code(601)
        empresa_b = self.empresa_service.get_empresa_by_code(602)
        documentos_a = self.documento_service.list_documentos_empresa(empresa_a["id"])
        documentos_b = self.documento_service.list_documentos_empresa(empresa_b["id"])
        tipo_balancetes = self.tipo_service.get_or_create_tipo("Balancetes")
        periodos = self.periodo_service.list_periodos()
        janeiro = next(item for item in periodos if item["ano"] == 2026 and item["mes"] == 1)
        marco = next(item for item in periodos if item["ano"] == 2026 and item["mes"] == 3)
        controle = self.status_service.build_control_view(empresa_a["id"], janeiro["id"], marco["id"])
        banco_a = next(
            doc
            for group in controle["groups"]
            for doc in group["documentos"]
            if doc["nome_documento"] == "Banco A"
        )

        self.assertEqual(result["processed_rows"], 3)
        self.assertEqual(result["companies_created"], 2)
        self.assertEqual(result["companies_updated"], 0)
        self.assertEqual(result["companies_reused"], 0)
        self.assertEqual(result["types_created"], 1)
        self.assertEqual(result["documents_imported"], 2)
        self.assertEqual(result["statuses_imported"], 3)
        self.assertEqual(result["failed"], 0)
        self.assertEqual(empresa_a["email_contato"], "empresa.a@teste.com")
        self.assertEqual(empresa_a["observacao"], "Observacao A")
        self.assertEqual(empresa_b["meios_recebimento"], None)
        self.assertEqual(
            [(doc["nome_documento"], doc["meios_recebimento"], doc["nome_tipo"]) for doc in documentos_a],
            [("Relatorio Gerencial", "Portal", "Balancetes"), ("Banco A", "Email", "Extratos CC")],
        )
        self.assertEqual(documentos_b, [])
        self.assertEqual([cell["status"] for cell in banco_a["cells"]], ["Recebido", "Pendente", "Nao cobrar"])
        self.assertEqual(tipo_balancetes["nome_tipo"], "Balancetes")

    def test_import_service_complete_import_updates_existing_company_and_rolls_back_failed_rows(self) -> None:
        class FakeWorksheet:
            def __init__(self, rows):
                self._rows = rows

            def iter_rows(self, min_row=1, values_only=True):
                return iter(self._rows)

        class FakeWorkbook:
            def __init__(self, rows):
                self.active = FakeWorksheet(rows)

        self.empresa_service.create_empresa(610, "Empresa Existente")
        workbook = FakeWorkbook(
            [
                (
                    "codigo_empresa",
                    "nome_empresa",
                    "email_contato",
                    "nome_contato",
                    "meios_recebimento",
                    "nome_documento",
                    "nome_tipo",
                    "observacao",
                ),
                (
                    610,
                    "Empresa Existente Atualizada",
                    "empresa.existente@teste.com",
                    "Carla",
                    "Email, Onvio",
                    "Banco Central",
                    "Extratos CC",
                    "Atualizada via importacao completa",
                ),
                (
                    611,
                    "Empresa Com Falha",
                    "empresa.falha@teste.com",
                    "Diego",
                    "Email",
                    "Documento sem tipo",
                    None,
                    "Nao deve persistir",
                ),
            ]
        )

        with patch.object(self.import_service, "_load_workbook", return_value=workbook):
            result = self.import_service.import_cadastros_completos("cadastro_completo_erros.xlsx")

        empresa_atualizada = self.empresa_service.get_empresa_by_code(610)
        documentos = self.documento_service.list_documentos_empresa(empresa_atualizada["id"])

        self.assertEqual(result["processed_rows"], 1)
        self.assertEqual(result["companies_created"], 0)
        self.assertEqual(result["companies_updated"], 1)
        self.assertEqual(result["documents_imported"], 1)
        self.assertEqual(result["statuses_imported"], 0)
        self.assertEqual(result["failed"], 1)
        self.assertIsNone(self.empresa_service.get_empresa_by_code(611))
        self.assertEqual(empresa_atualizada["nome_empresa"], "Empresa Existente Atualizada")
        self.assertIsNone(empresa_atualizada["meios_recebimento"])
        self.assertEqual(empresa_atualizada["email_contato"], "empresa.existente@teste.com")
        self.assertEqual(empresa_atualizada["nome_contato"], "Carla")
        self.assertEqual(empresa_atualizada["observacao"], "Atualizada via importacao completa")
        self.assertEqual([doc["nome_documento"] for doc in documentos], ["Banco Central"])
        self.assertEqual(documentos[0]["meios_recebimento"], "Email, Onvio")
        self.assertIn("Linha 3", result["errors"][0])

    def test_import_service_complete_import_ignores_month_columns_on_company_only_rows(self) -> None:
        class FakeWorksheet:
            def __init__(self, rows):
                self._rows = rows

            def iter_rows(self, min_row=1, values_only=True):
                return iter(self._rows)

        class FakeWorkbook:
            def __init__(self, rows):
                self.active = FakeWorksheet(rows)

        self.periodo_service.generate_year(2026)
        workbook = FakeWorkbook(
            [
                (
                    "codigo_empresa",
                    "nome_empresa",
                    "email_contato",
                    "nome_contato",
                    "nome_documento",
                    "meio de recebimento",
                    "tipos",
                    "observacoes",
                    "janeiro",
                ),
                (
                    612,
                    "Empresa So Cadastro",
                    "empresa@teste.com",
                    "Nadia",
                    None,
                    None,
                    None,
                    "Sem documentos ainda",
                    "email-indevido@teste.com",
                ),
            ]
        )

        with patch.object(self.import_service, "_load_workbook", return_value=workbook):
            result = self.import_service.import_cadastros_completos("cadastro_completo_so_empresa.xlsx")

        empresa = self.empresa_service.get_empresa_by_code(612)
        documentos = self.documento_service.list_documentos_empresa(empresa["id"])

        self.assertEqual(result["processed_rows"], 1)
        self.assertEqual(result["documents_imported"], 0)
        self.assertEqual(result["statuses_imported"], 0)
        self.assertEqual(result["failed"], 0)
        self.assertEqual(empresa["observacao"], "Sem documentos ainda")
        self.assertEqual(documentos, [])

    def test_alias_type_name_is_normalized_to_existing_canonical_type(self) -> None:
        tipo = self.tipo_service.get_or_create_tipo("Extrato CC")

        self.assertEqual(tipo["nome_tipo"], "Extratos CC")

    def test_duplicate_type_aliases_are_merged_without_losing_status(self) -> None:
        empresa_id = self.empresa_service.create_empresa(501, "Empresa Alias")
        canonical_type_id = self.tipo_service.get_or_create_tipo("Extratos CC")["id"]
        alias_type_id = self.tipo_repository.create("Extrato CC", TYPE_OCCURRENCE_MENSAL)

        canonical_document_id = self.documento_repository.create(empresa_id, canonical_type_id, None, "Banco XP")
        alias_document_id = self.documento_repository.create(empresa_id, alias_type_id, None, "Banco XP")

        self.periodo_service.generate_year(2026)
        periodos = self.periodo_service.list_periodos()
        jan = next(item for item in periodos if item["ano"] == 2026 and item["mes"] == 1)
        feb = next(item for item in periodos if item["ano"] == 2026 and item["mes"] == 2)

        self.status_service.update_status(canonical_document_id, jan["id"], "Recebido")
        self.status_service.update_status(alias_document_id, feb["id"], "Pendente")

        initialize_schema(self.empresa_repository.db_manager)

        tipos = self.tipo_service.list_tipos()
        nomes = [tipo["nome_tipo"] for tipo in tipos]
        self.assertEqual(nomes.count("Extratos CC"), 1)
        self.assertNotIn("Extrato CC", nomes)

        documentos = self.documento_service.list_documentos_empresa(empresa_id)
        self.assertEqual(len([doc for doc in documentos if doc["nome_documento"] == "Banco XP"]), 1)

        view = self.status_service.build_control_view(empresa_id, jan["id"], feb["id"])
        self.assertEqual(len(view["groups"]), 1)
        self.assertEqual(view["groups"][0]["tipo_nome"], "Extratos CC")
        self.assertEqual(view["groups"][0]["documentos"][0]["cells"][0]["status"], "Recebido")
        self.assertEqual(view["groups"][0]["documentos"][0]["cells"][1]["status"], "Pendente")

    def test_default_admin_user_can_authenticate(self) -> None:
        user = self.auth_service.authenticate("admin", "admin")

        self.assertEqual(user["username"], "admin")
        self.assertEqual(user["tipo_usuario"], "admin")
        self.assertEqual(user["ativa"], 1)

    def test_auth_service_can_authenticate_with_remembered_session(self) -> None:
        token = self.auth_service.create_remembered_session(self.session_service.get_user_id())

        user, refreshed_token = self.auth_service.authenticate_with_remembered_session(token)

        self.assertEqual(user["username"], "admin")
        self.assertEqual(user["tipo_usuario"], "admin")
        self.assertNotEqual(token, refreshed_token)

        with self.assertRaises(ValidationError):
            self.auth_service.authenticate_with_remembered_session(token)

    def test_expired_remembered_session_cannot_authenticate(self) -> None:
        token = self.auth_service.create_remembered_session(self.session_service.get_user_id())
        selector = token.split(".", 1)[0]
        expired_timestamp = (datetime.now(timezone.utc) - timedelta(days=61)).strftime("%Y-%m-%d %H:%M:%S")

        with self.remembered_session_repository.db_manager.connect() as connection:
            connection.execute(
                """
                UPDATE sessoes_lembradas
                SET criado_em = ?, ultimo_uso_em = ?
                WHERE selector = ?
                """,
                (expired_timestamp, expired_timestamp, selector),
            )

        with self.assertRaises(ValidationError):
            self.auth_service.authenticate_with_remembered_session(token)

    def test_revoked_remembered_session_cannot_authenticate(self) -> None:
        token = self.auth_service.create_remembered_session(self.session_service.get_user_id())

        self.auth_service.revoke_remembered_session(token)

        with self.assertRaises(ValidationError):
            self.auth_service.authenticate_with_remembered_session(token)

    def test_password_change_revokes_user_remembered_sessions(self) -> None:
        user_id = self.user_service.create_user("lembrado", "123456", "comum", ativo=True)
        token = self.auth_service.create_remembered_session(user_id)

        self.user_service.update_user(user_id, "lembrado", "comum", True, password="654321")

        with self.assertRaises(ValidationError):
            self.auth_service.authenticate_with_remembered_session(token)

    def test_restore_backup_rejects_incompatible_sqlite_database(self) -> None:
        invalid_backup = Path(self.temp_dir.name) / "outro_sqlite.db"

        import sqlite3

        with sqlite3.connect(invalid_backup) as connection:
            connection.execute("CREATE TABLE exemplo (id INTEGER PRIMARY KEY)")

        with self.assertRaises(ValidationError):
            self.database_maintenance_service.restore_backup(invalid_backup)

    def test_inactive_user_cannot_login(self) -> None:
        user_id = self.user_service.create_user("maria", "123456", "comum", ativo=True)
        self.user_service.update_user(user_id, "maria", "comum", False)

        with self.assertRaises(ValidationError):
            self.auth_service.authenticate("maria", "123456")

    def test_logs_are_created_for_empresa_documento_e_status(self) -> None:
        empresa_id = self.empresa_service.create_empresa(777, "HND TEC")
        tipo_id = self.tipo_service.get_or_create_tipo("Extratos CC")["id"]
        documento_id = self.documento_service.create_documento(empresa_id, tipo_id, "Banco do Brasil")
        self.periodo_service.generate_year(2026)
        periodo = next(
            item for item in self.periodo_service.list_periodos() if item["ano"] == 2026 and item["mes"] == 3
        )

        self.status_service.update_status(documento_id, periodo["id"], "Recebido")

        logs = self.log_service.list_logs(limit=20)
        actions = [log["acao"] for log in logs]
        descriptions = [log["descricao"] for log in logs]

        self.assertIn("CADASTRO_EMPRESA", actions)
        self.assertIn("CADASTRO_DOCUMENTO", actions)
        self.assertIn("ALTERACAO_STATUS", actions)
        self.assertTrue(any("Banco do Brasil" in description for description in descriptions))
        self.assertTrue(any("HND TEC" in description for description in descriptions))

    def test_status_closure_logs_removed_future_statuses(self) -> None:
        empresa_id = self.empresa_service.create_empresa(778, "Empresa Encerramento")
        tipo_id = self.tipo_service.get_or_create_tipo("Extratos CC")["id"]
        documento_id = self.documento_service.create_documento(empresa_id, tipo_id, "Banco Futuro")
        self.periodo_service.generate_year(2026)
        periodos = self.periodo_service.list_periodos()
        abril = next(item for item in periodos if item["ano"] == 2026 and item["mes"] == 4)
        junho = next(item for item in periodos if item["ano"] == 2026 and item["mes"] == 6)

        self.status_service.update_status(documento_id, junho["id"], "Recebido")
        self.status_service.update_status(documento_id, abril["id"], "Encerrado")

        logs = self.log_service.list_logs(limit=20, empresa_id=empresa_id)
        future_logs = [log for log in logs if log["periodo_ano"] == 2026 and log["periodo_mes"] == 6]

        self.assertTrue(future_logs)
        self.assertTrue(any('de "Recebido" para "vazio"' in log["descricao"] for log in future_logs))

    def test_company_creation_rolls_back_if_audit_log_fails(self) -> None:
        with patch.object(self.audit_service, "log", side_effect=RuntimeError("falha no log")):
            with self.assertRaises(RuntimeError):
                self.empresa_service.create_empresa(779, "Empresa Sem Log")

        self.assertIsNone(self.empresa_service.get_empresa_by_code(779))

    def test_log_filters_by_empresa_e_periodo(self) -> None:
        empresa_a_id = self.empresa_service.create_empresa(801, "Empresa A")
        empresa_b_id = self.empresa_service.create_empresa(802, "Empresa B")
        tipo_id = self.tipo_service.get_or_create_tipo("Extratos CC")["id"]
        documento_a_id = self.documento_service.create_documento(empresa_a_id, tipo_id, "Doc A")
        documento_b_id = self.documento_service.create_documento(empresa_b_id, tipo_id, "Doc B")

        self.periodo_service.generate_year(2026)
        periodos = self.periodo_service.list_periodos()
        jan = next(item for item in periodos if item["ano"] == 2026 and item["mes"] == 1)
        feb = next(item for item in periodos if item["ano"] == 2026 and item["mes"] == 2)

        self.status_service.update_status(documento_a_id, jan["id"], "Recebido")
        self.status_service.update_status(documento_b_id, feb["id"], "Pendente")

        company_logs = self.log_service.list_logs(limit=50, empresa_id=empresa_a_id)
        self.assertTrue(company_logs)
        self.assertTrue(all(log["empresa_id"] == empresa_a_id for log in company_logs))

        period_logs = self.log_service.list_logs(limit=50, periodo_ano=2026, periodo_mes=2)
        self.assertTrue(period_logs)
        self.assertTrue(all(log["periodo_ano"] == 2026 and log["periodo_mes"] == 2 for log in period_logs))
        self.assertTrue(any("Empresa B" in log["descricao"] for log in period_logs))


if __name__ == "__main__":
    unittest.main()
